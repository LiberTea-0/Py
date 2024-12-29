import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side


class Part:
    def __init__(self, name, price, number, description):
        self.name = name
        self.price = price
        self.number = number
        self.description = description

class Project:
    def __init__(self, name):
        self.name = name
        self.parts = []

    def add_part(self, name, price, number, description):
        new_part = Part(name, price, number, description)
        self.parts.append(new_part)

    def delet_part(self, part_number):
        if 0 < part_number <= len(self.parts):
            delet_part = self.parts.pop(part_number - 1)
            print(f"Part '{delet_part.name}' has been deleted.")
        else:
            print("Invalid number!")

    def parts_list(self):
        if not self.parts:
            print(f"No part in the project '{self.name}'.")
        else:
            for i, part in enumerate(self.parts, start = 1):
                print(f"{i}. Name:{part.name}, price:{part.price}, number: {part.number}, description:{part.description}")
    
        
class System:
    def __init__(self):
        self.project = {}
        self.current_file = None
    
    def create_project(self, project_name):
        if project_name in self.project:
            print("The project is already exist")
        else:
            self.project[project_name] = Project(project_name)
            print(f"\"{project_name}\" is successfully created")
    
    def projects_list(self):
        if not self.project:
            print("The project list is empty")
        else:
            for i, project_name in enumerate(self.project.keys(), start=1):
                print(f"{i}.name: {project_name}")

    def delet_project(self, delet_number):
        if not self.project:
            print("The project list is empty")
        else:
            project_name = list(self.project.keys())[delet_number - 1]
            del self.project[project_name]
            print(f"Project '{project_name}' deleted successfully!")


    def save_to_excel(self, filename):
        if not self.current_file:
            filename = input("Enter the filename to save: ")
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            self.current_file = filename

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Projects and Parts"

        header = ["Project Name", "Part Name", "Price", "Number", "Description"]
        sheet.append(header)

        header_font = Font(bold=True, color="FFFFFF")  
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")  
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col_num, cell in enumerate(sheet[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for project_name, project in self.project.items():
            for i, part in enumerate(project.parts):
                if i == 0:
                    sheet.append([project_name, part.name, part.price, part.number, part.description])
                else:
                    sheet.append([None, part.name, part.price, part.number, part.description])    


        for row in sheet.iter_rows(min_row=2, values_only=False):  
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
        
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter 
            for cell in column_cells:
                try:
                    if cell.value:  
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  
            sheet.column_dimensions[column].width = adjusted_width
        
        try:
            workbook.save(self.current_file)
            print(f"Data successfully saved to {filename}")
        except Exception as e:
            print(f"Failed to save Excel file: {e}")

    def load_from_excel(self, filename):
        if not os.path.exists(filename):
            print(f"File {filename} does not exit. Starting with an empty project.")
            return

        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active

            current_project_name = None

            for row in sheet.iter_rows(min_row=2, values_only=True):
                project_name, part_name, price, number, description = row or (None,None,None,None,None)                           
                
                if project_name:
                    current_project_name = project_name
                
                if not current_project_name:
                    continue

                if current_project_name not in self.project: #缺失后同project非首排part会因project名为None而load失败
                    self.project[current_project_name] = Project(current_project_name)

                self.project[current_project_name].add_part(part_name, price, number, description)
            
            self.current_file = filename

            print(f"Data successfully loaded from {filename}")
        except Exception as e:
            print(f"Failed to load Excel file: {e}")

    def select_file(self):
        excel_files = [f for f in os.listdir() if f.endswith(".xlsx")]

        if excel_files:
            print("Excel list in the prescent contents")
            for i, file in enumerate(excel_files, 1):
                print(f"{i}. {file}")

            choice = input("Choose the file number: ")

            try:
                choice = int(choice)
                if 1 <= choice <= len(excel_files):
                    excel_file = excel_files[choice - 1]
                else:
                    print("Error Number")
                    return None

            except ValueError:
                    print("Error, Exit!")
                    return None
        else:

            excel_file = input("No Excel file, enter any key to jump. ")
            if not excel_file.endswith(".xlsx"):
                excel_file += ".xlsx"
        
        self.current_file = excel_file

        return excel_file

    def Menu_Layer1(self):
        print("\n Parts manage system")
        print("1、Show the prescent project")
        print("2、New a project")
        print("3、Enter a project")
        print("4、Delet project")
        print("5、Save and exit")

    def Menu_Layer2(self):
        print("\n1、Add Parts")
        print("2、Show parts list")
        print("3、Back to the main menu")
        print("4、Delet Parts")
    
def main():
    system = System()
    excel_file = system.select_file()
    if excel_file is None:
        return
    system.load_from_excel(excel_file)

    while True:
        
        system.Menu_Layer1()
        choice = input("\nEnter your choice: ")

        if choice == "1":
            system.projects_list()
        elif choice == "2":
            project_name = input("Name the project: ")
            system.create_project(project_name)
        elif choice == "3":
            system.projects_list()
            try:
                project_number = int(input("Input the project number: "))
                if project_number < 1 or project_number > len(system.project):
                    print("Input wrong!")
                else:
                    project_name = list(system.project.keys())[project_number - 1]
                    project = system.project[project_name]
                    project.parts_list()
            except ValueError:
                print("Invalid input! Enter an number. ")
                
            while True:

                    system.Menu_Layer2()
                    sub_choice = input("Enter your choice: ")

                    if sub_choice == "1":
                        part_name = input("Part name: ")
                        try:
                            part_price = float(input("Part price: "))
                        except ValueError:
                            print("Invalid price input. Enter a number!")
                            continue
                            
                        try:
                            part_number = int(input("number: "))
                        except ValueError:
                            print("Invalid input. Enter a number")
                            continue
                        part_description = input("Description: ")

                        project.add_part(part_name, part_price, part_number, part_description)
                        print("Add successfully!")
                    elif sub_choice == "2":
                        project.parts_list()
                    elif sub_choice == "3":
                        break
                    elif sub_choice == "4":
                        project.parts_list()
                        try:
                            target = int(input("Input the target item number and delet it: "))
                            project.delet_part(target)
                        except ValueError:
                            print("Invalid input. Please enter a valid number!")


                    else:
                        print("Wrong operation, try it again!")
        elif choice == "4":
            system.projects_list()          
            try:
                delet_number = int(input("Select the number: "))
                if delet_number < 1 or delet_number > len(system.project):
                    print("Invalid number.")
                else:
                    system.delet_project(delet_number)
            except ValueError:
                print("Invalid Input.")

        elif choice == "5":
            if system.current_file: 
                system.save_to_excel(system.current_file)
                print("Data successfully saved!")
            else: 
                filename = input("Enter the filename to save before exiting: ")
                if not filename.endswith(".xlsx"):
                    filename += ".xlsx"
                system.save_to_excel(filename)
                print("Data successfully saved!")
            break


if __name__ == "__main__":
    main()
